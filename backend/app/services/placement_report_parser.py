"""
Парсер отчёта seller_placement_by_products из Ozon Report API.

Структура (per-day-per-SKU-per-warehouse, 12 колонок):
  Дата | SKU | Артикул | Категория товара | Описательный тип | Склад |
  Признак товара | Суммарный объем в мл | Кол-во экземпляров |
  Платный объем в мл | Кол-во платных экземпляров | Начисленная стоимость размещения

Отчёт даёт ТОЛЬКО хранение (storage) — реклама, эквайринг, комиссия в нём нет.
Это идеальный источник для поля monthly_unit_economy.storage:
агрегируем SUM(storage_cost) per (sku, month).
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any, BinaryIO

from openpyxl import load_workbook

# Возможные написания колонок (Ozon иногда меняет)
HEADER_DATE = {"Дата", "Date"}
HEADER_SKU = {"SKU"}
HEADER_OFFER = {"Артикул", "Offer"}
HEADER_WAREHOUSE = {"Склад", "Warehouse"}
HEADER_STORAGE_COST = {"Начисленная стоимость размещения", "Storage cost"}


@dataclass
class PlacementDailyRow:
    """Одна сырая daily-запись из API-отчёта."""
    sku: int
    warehouse: str
    day: date
    storage_cost: Decimal  # знак как в файле (мы инвертируем при записи)
    offer_id: str | None = None


@dataclass
class PlacementParseResult:
    # Сырые daily-строки — пишем как есть в placement_storage_daily
    # с PK (cabinet_id, sku, warehouse, day) → перекрытие отчётов
    # дедуплицируется естественно.
    daily_rows: list[PlacementDailyRow] = field(default_factory=list)
    # Метаданные
    offer_by_sku: dict[int, str] = field(default_factory=dict)
    rows_total: int = 0
    rows_with_cost: int = 0
    period_from: date | None = None
    period_to: date | None = None


def _norm(s: Any) -> str:
    return re.sub(r"\s+", " ", str(s or "")).strip()


def _to_decimal(v: Any) -> Decimal | None:
    if v is None or v == "":
        return None
    try:
        if isinstance(v, (int, float, Decimal)):
            return Decimal(str(v))
        s = str(v).replace(" ", "").replace("\xa0", "")
        if "," in s and "." not in s:
            s = s.replace(",", ".")
        elif "," in s and "." in s:
            s = s.replace(",", "")
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def parse_placement_report(file_obj: BinaryIO) -> PlacementParseResult:
    """
    Возвращает агрегат storage per (sku, month).

    Бросает ValueError если ключевые колонки (Дата, SKU, Стоимость размещения)
    не найдены в заголовке.
    """
    wb = load_workbook(file_obj, read_only=True, data_only=True)
    # Используем первый лист (обычно «Страница #1»)
    ws = wb[wb.sheetnames[0]]

    # Заголовки в строке 1
    headers_raw = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    headers = [_norm(h) for h in headers_raw]

    def find_col(names: set[str]) -> int | None:
        for i, h in enumerate(headers):
            if h in names:
                return i
        return None

    col_date = find_col(HEADER_DATE)
    col_sku = find_col(HEADER_SKU)
    col_offer = find_col(HEADER_OFFER)
    col_warehouse = find_col(HEADER_WAREHOUSE)
    col_cost = find_col(HEADER_STORAGE_COST)

    missing = []
    if col_date is None: missing.append("Дата")
    if col_sku is None: missing.append("SKU")
    if col_cost is None: missing.append("Начисленная стоимость размещения")
    if missing:
        raise ValueError(
            f"Не найдены обязательные колонки: {', '.join(missing)}. "
            f"Получено: {headers[:15]}"
        )

    result = PlacementParseResult()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(c is None for c in row):
            continue
        result.rows_total += 1

        # Дата
        date_val = row[col_date]
        if isinstance(date_val, datetime):
            d = date_val.date()
        elif isinstance(date_val, date):
            d = date_val
        else:
            d = None
            if date_val:
                try:
                    d = datetime.fromisoformat(str(date_val).split(" ")[0]).date()
                except Exception:
                    pass
        if not d:
            continue

        # Обновляем period bounds
        if not result.period_from or d < result.period_from:
            result.period_from = d
        if not result.period_to or d > result.period_to:
            result.period_to = d

        # SKU
        sku_raw = row[col_sku]
        try:
            sku = int(sku_raw) if sku_raw else 0
        except (TypeError, ValueError):
            continue
        if not sku:
            continue

        # offer_id для отображения
        offer_id = None
        if col_offer is not None:
            offer_id = _norm(row[col_offer]) or None
            if offer_id:
                result.offer_by_sku.setdefault(sku, offer_id)

        # Склад — часть PK, не теряем (один SKU может быть на нескольких складах в один день)
        warehouse = ""
        if col_warehouse is not None:
            warehouse = _norm(row[col_warehouse])

        # Storage cost — сырое значение, не агрегируем (агрегацию делает SQL)
        cost = _to_decimal(row[col_cost])
        if cost is None:
            continue
        if cost != 0:
            result.rows_with_cost += 1

        result.daily_rows.append(PlacementDailyRow(
            sku=sku,
            warehouse=warehouse,
            day=d,
            storage_cost=cost,
            offer_id=offer_id,
        ))

    wb.close()
    return result
