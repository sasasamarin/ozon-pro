"""
Парсер отчёта о реализации Ozon (XLSX).

Реализация AUDIT.md A4 (раньше — заглушка).

Структура XLSX (фактически из /v2/finance/realization):
  - Лист 1 содержит шапку («Отчёт о реализации товаров за <месяц>»)
  - Затем таблица с колонками:
    * № строки
    * Артикул / Название товара
    * Цена продавца за единицу
    * Доставка покупателю (количество, цена, сумма, бонус, комиссия, итого)
    * Возврат (количество, цена, сумма, бонус, комиссия, итого)
  - Внизу — totals row

openpyxl читает в data_only=True (без формул). Парсер агрегирует
числовые колонки с эвристикой по именам.
"""
from __future__ import annotations

import io
from dataclasses import dataclass, field
from typing import Any

import openpyxl

from app.core.logging import log


@dataclass
class RealizationRow:
    sku: int | None = None
    name: str | None = None
    seller_price: float | None = None
    delivered_qty: int = 0
    delivered_amount: float = 0.0
    delivered_bonus: float = 0.0       # СПП-компенсация
    delivered_commission: float = 0.0
    returned_qty: int = 0
    returned_amount: float = 0.0


@dataclass
class OzonRealizationReport:
    revenue: float | None = None           # Σ delivered_amount
    commission: float | None = None        # Σ delivered_commission
    bonus_compensation: float | None = None  # Σ delivered_bonus (СПП от Ozon)
    payout: float | None = None            # revenue + bonus - commission - returns
    delivered_qty: int = 0
    returned_qty: int = 0
    rows: list[RealizationRow] = field(default_factory=list)
    period_label: str | None = None        # человеческий «Март 2026»
    raw_rows: list[dict] = field(default_factory=list)


def _normalize(s: Any) -> str:
    if s is None:
        return ""
    return " ".join(str(s).split()).lower()


def _find_col(headers: list[str], needles: list[str]) -> int | None:
    """Найти первую колонку, чей нормализованный заголовок содержит любую из подстрок."""
    for i, h in enumerate(headers):
        h_low = _normalize(h)
        for n in needles:
            if n in h_low:
                return i
    return None


def _to_float(v: Any) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        s = str(v).replace(",", ".").replace(" ", "").replace("\xa0", "")
        return float(s)
    except (ValueError, TypeError):
        return 0.0


def _to_int(v: Any) -> int:
    try:
        return int(_to_float(v))
    except (ValueError, TypeError):
        return 0


def _parse_workbook(wb: openpyxl.Workbook) -> OzonRealizationReport:
    sheet = wb.active
    if sheet is None:
        return OzonRealizationReport()

    # Тянем первые 200 строк — достаточно для шапки + таблицы среднего отчёта.
    # Большие отчёты могут иметь тысячи строк — но мы парсим всё, не первые 200.
    all_rows = list(sheet.iter_rows(values_only=True))
    if not all_rows:
        return OzonRealizationReport()

    # 1) Найти период: ищем строку с словом «период» / «отчёт»
    period_label = None
    for r in all_rows[:10]:
        if not r:
            continue
        for cell in r:
            if cell and isinstance(cell, str) and ("отчёт" in cell.lower() or "период" in cell.lower()):
                period_label = cell.strip()
                break
        if period_label:
            break

    # 2) Найти строку-заголовок таблицы — обычно содержит «sku» или «артикул»
    header_idx = None
    for i, r in enumerate(all_rows):
        if not r:
            continue
        if any(c and isinstance(c, str) and ("артикул" in c.lower() or "товар" in c.lower())
               for c in r):
            header_idx = i
            break
    if header_idx is None:
        log.warning("realization_no_header")
        return OzonRealizationReport(period_label=period_label)

    headers = [str(h or "").strip() for h in all_rows[header_idx]]

    # 3) Найти ключевые колонки (эвристика по подстрокам)
    sku_col = _find_col(headers, ["sku", "артикул"])
    name_col = _find_col(headers, ["товар", "название"])
    price_col = _find_col(headers, ["цена продавца", "продавца за"])
    # Delivery
    deliv_qty_col = _find_col(headers, ["доставка покупателю.*количество", "продано", "доставка количество"])
    if deliv_qty_col is None:
        deliv_qty_col = _find_col(headers, ["количество"])
    deliv_amount_col = _find_col(headers, ["доставка покупателю.*сумма", "сумма доставки", "к выплате"])
    if deliv_amount_col is None:
        deliv_amount_col = _find_col(headers, ["сумма"])
    deliv_bonus_col = _find_col(headers, ["бонус", "баллы", "спп"])
    deliv_comm_col = _find_col(headers, ["комиссия"])
    # Returns
    ret_qty_col = _find_col(headers, ["возврат.*количество", "возвращ"])
    ret_amount_col = _find_col(headers, ["возврат.*сумма", "сумма возврата"])

    rows: list[RealizationRow] = []
    raw_rows: list[dict] = []

    for row in all_rows[header_idx + 1:]:
        if not row or not any(row):
            continue
        # Защита от итоговой строки «Итого» — обычно sku пуст / название «Итого»
        sku_val = row[sku_col] if sku_col is not None and sku_col < len(row) else None
        name_val = row[name_col] if name_col is not None and name_col < len(row) else None
        if name_val and "итого" in _normalize(name_val):
            continue

        rr = RealizationRow(
            sku=_to_int(sku_val) if sku_val else None,
            name=str(name_val) if name_val else None,
            seller_price=_to_float(row[price_col]) if price_col is not None and price_col < len(row) else None,
            delivered_qty=_to_int(row[deliv_qty_col]) if deliv_qty_col is not None and deliv_qty_col < len(row) else 0,
            delivered_amount=_to_float(row[deliv_amount_col]) if deliv_amount_col is not None and deliv_amount_col < len(row) else 0.0,
            delivered_bonus=_to_float(row[deliv_bonus_col]) if deliv_bonus_col is not None and deliv_bonus_col < len(row) else 0.0,
            delivered_commission=_to_float(row[deliv_comm_col]) if deliv_comm_col is not None and deliv_comm_col < len(row) else 0.0,
            returned_qty=_to_int(row[ret_qty_col]) if ret_qty_col is not None and ret_qty_col < len(row) else 0,
            returned_amount=_to_float(row[ret_amount_col]) if ret_amount_col is not None and ret_amount_col < len(row) else 0.0,
        )
        # Пропускаем строки без данных (только заголовочные значения)
        if rr.sku is None and not rr.name and not rr.delivered_amount:
            continue
        rows.append(rr)

        raw_rows.append({h: row[i] for i, h in enumerate(headers)
                         if i < len(row) and row[i] is not None})

    # Агрегаты
    total_rev = sum(r.delivered_amount for r in rows)
    total_comm = sum(r.delivered_commission for r in rows)
    total_bonus = sum(r.delivered_bonus for r in rows)
    total_ret = sum(r.returned_amount for r in rows)
    total_deliv_qty = sum(r.delivered_qty for r in rows)
    total_ret_qty = sum(r.returned_qty for r in rows)

    return OzonRealizationReport(
        revenue=round(total_rev, 2),
        commission=round(total_comm, 2),
        bonus_compensation=round(total_bonus, 2),
        # Payout = revenue + bonus_от_Ozon − commission − returns
        payout=round(total_rev + total_bonus - total_comm - total_ret, 2),
        delivered_qty=total_deliv_qty,
        returned_qty=total_ret_qty,
        rows=rows,
        period_label=period_label,
        raw_rows=raw_rows[:50],  # ограничиваем для размера ответа
    )


def parse_realization_xlsx(file_path_or_bytes: str | bytes) -> OzonRealizationReport:
    """
    Принимает путь к файлу ИЛИ байты XLSX (из upload).
    Возвращает агрегированный отчёт.
    """
    try:
        if isinstance(file_path_or_bytes, bytes):
            wb = openpyxl.load_workbook(
                io.BytesIO(file_path_or_bytes), read_only=True, data_only=True,
            )
        else:
            wb = openpyxl.load_workbook(file_path_or_bytes, read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001
        log.exception("realization_load_failed", err=str(e))
        return OzonRealizationReport()

    return _parse_workbook(wb)
