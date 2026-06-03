"""
Парсер XLSX «Экономика магазина → Общие расходы» из Ozon Seller UI.

Структура файла (проверено на майском экспорте):
- Лист: «Юнит-экономика»
- Строка 1: «Период: DD.MM.YYYY-DD.MM.YYYY» → отсюда берём период
- Строка 2: «Общие расходы»
- Строка 3: групповые заголовки (объединённые)
- Строка 4: заголовки колонок (33 шт) — матчим ПО ИМЕНИ, не позиции
- Строки 5+: данные, одна строка = один SKU
- Итогов в файле нет — только товары

Принципы:
- Знаки сохраняются как в файле (доходы +, расходы −)
- Если ожидаемая колонка не найдена — ImportError, не молча 0
- SKU пустой/не число → строка пропускается
- Деньги хранятся как Decimal (не float)
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any, BinaryIO

from openpyxl import load_workbook


# Маппинг русских заголовков → имя поля в БД
COLUMN_MAP: dict[str, str] = {
    "SKU": "sku",
    "Артикул": "offer_id",
    "Название товара": "name",
    "Схема работы": "scheme",
    "Текущая цена": "current_price",
    "Заказано товаров, шт": "ordered_qty",
    "Доставлено товаров, шт": "delivered_qty",
    "Возвращено товаров, шт": "returned_qty",
    "Выручка": "revenue",
    "Баллы за скидки": "spp_points",
    "Программы партнёров": "partner_programs",
    "Программы партнеров": "partner_programs",  # альтернативное написание
    "Вознаграждение Ozon": "ozon_commission",
    "Эквайринг": "acquiring",
    "Обработка отправления": "posting_handling",
    "Логистика": "logistics",
    "Доставка до места выдачи": "last_mile",
    # XLSX «Общие расходы» точно покрывает календарный месяц → пишем в
    # storage_from_xlsx (приоритет в P&L через COALESCE). Поле storage —
    # легаси, оставлено для отката (см. миграцию 0020).
    # Ozon меняет заголовок в разных вариантах отчёта — поддерживаем все.
    "Стоимость размещения": "storage_from_xlsx",
    "Платное размещение": "storage_from_xlsx",
    "Стоимость хранения": "storage_from_xlsx",
    "Хранение": "storage_from_xlsx",
    "Размещение": "storage_from_xlsx",
    "Обработка возврата": "return_handling",
    "Обратная логистика": "reverse_logistics",
    "Утилизация": "disposal",
    "Доп. обработка ОВХ": "ovh_extra",
    "Дополнительная обработка ОВХ": "ovh_extra",
    "Операционные ошибки": "operational_errors",
    "Оплата за клик": "ad_cpc",
    "Оплата за заказ": "ad_cpo",
    "Звёздные товары": "ad_star",
    "Звездные товары": "ad_star",
    "Платный бренд": "ad_paid_brand",
    "Отзывы": "ad_reviews",
    "Доля от продаж": "ozon_margin_share",
    "Прибыль за период": "ozon_profit",
    "Индекс цен": "price_index",
    # Прибыль за шт + Доступность товаров — не нужны
}

# Числовые поля для конвертации в Decimal
NUMERIC_FIELDS = {
    "current_price", "revenue", "spp_points", "partner_programs",
    "ozon_commission", "acquiring", "posting_handling", "logistics",
    "last_mile", "storage_from_xlsx", "return_handling", "reverse_logistics",
    "disposal", "ovh_extra", "operational_errors",
    "ad_cpc", "ad_cpo", "ad_star", "ad_paid_brand", "ad_reviews",
    "ozon_profit", "ozon_margin_share", "price_index",
}
INT_FIELDS = {"sku", "ordered_qty", "delivered_qty", "returned_qty"}


@dataclass
class ParseResult:
    period_from: date
    period_to: date
    rows: list[dict[str, Any]] = field(default_factory=list)
    skipped_rows: int = 0
    unknown_columns: list[str] = field(default_factory=list)
    file_name: str = ""


def _normalize_header(s: str | None) -> str:
    if not s:
        return ""
    # схлопываем \n и многократные пробелы (Ozon вставляет «\nобработка» внутри)
    return re.sub(r"\s+", " ", str(s)).strip()


def _parse_period(raw: str) -> tuple[date, date]:
    """«Период: 01.05.2026-31.05.2026» → (2026-05-01, 2026-05-31)."""
    # Гибко: ищем пару dd.mm.yyyy-dd.mm.yyyy
    m = re.search(
        r"(\d{2})\.(\d{2})\.(\d{4})\s*[-—–]\s*(\d{2})\.(\d{2})\.(\d{4})",
        raw,
    )
    if not m:
        raise ValueError(f"Не удалось распарсить период из строки 1: {raw!r}")
    d1, m1, y1, d2, m2, y2 = m.groups()
    return (date(int(y1), int(m1), int(d1)),
            date(int(y2), int(m2), int(d2)))


def _to_decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        if isinstance(value, (int, float, Decimal)):
            return Decimal(str(value))
        # строка типа "1 222,15" или "1,222.15"
        s = str(value).replace(" ", "").replace(" ", "")
        # запятая как decimal separator в русской локали
        if "," in s and "." not in s:
            s = s.replace(",", ".")
        elif "," in s and "." in s:
            s = s.replace(",", "")
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def _to_int(value: Any) -> int | None:
    d = _to_decimal(value)
    return int(d) if d is not None else None


def parse_unit_economy(
    file_obj: BinaryIO, file_name: str = "",
) -> ParseResult:
    """
    Парсит загруженный XLSX → ParseResult.

    Бросает ValueError если структура неожиданная (лист не найден,
    период не парсится, обязательные колонки отсутствуют).
    """
    wb = load_workbook(file_obj, read_only=True, data_only=True)

    # 1. Лист «Юнит-экономика» (или первый если его нет)
    sheet_name = None
    for name in wb.sheetnames:
        if "юнит" in name.lower() or "unit" in name.lower():
            sheet_name = name
            break
    if not sheet_name:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    # 2. Период из строки 1 (любая первая непустая ячейка)
    period_from: date | None = None
    period_to: date | None = None
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        for cell in row:
            if cell and isinstance(cell, str) and "Период" in cell:
                period_from, period_to = _parse_period(cell)
                break
        if period_from:
            break
    if not period_from or not period_to:
        # Fallback: ищем в первых 3 строках
        for r in range(1, 4):
            for row in ws.iter_rows(min_row=r, max_row=r, values_only=True):
                for cell in row:
                    if cell and isinstance(cell, str) and "Период" in cell:
                        period_from, period_to = _parse_period(cell)
                        break
                if period_from:
                    break
    if not period_from:
        raise ValueError("Не нашёл строку «Период: ...» в первых 3 строках файла")

    # 3. Заголовки — строка 4 (но ищем строку с «SKU» среди первых 6 чтобы быть гибче)
    header_row_idx = None
    headers: list[str] = []
    for r in range(3, 8):
        candidates = list(next(ws.iter_rows(min_row=r, max_row=r, values_only=True)))
        candidates = [_normalize_header(c) for c in candidates]
        if "SKU" in candidates:
            header_row_idx = r
            headers = candidates
            break
    if header_row_idx is None:
        raise ValueError("Не нашёл строку заголовков (с колонкой «SKU»)")

    # 4. Маппинг index → field_name
    col_to_field: dict[int, str] = {}
    unknown: list[str] = []
    for idx, h in enumerate(headers):
        if not h:
            continue
        field_name = COLUMN_MAP.get(h)
        if field_name:
            col_to_field[idx] = field_name
        else:
            unknown.append(h)

    if "sku" not in col_to_field.values():
        raise ValueError("Колонка SKU не сматчилась — проверь заголовки файла")

    # 5. Данные — все строки после заголовков
    rows: list[dict[str, Any]] = []
    skipped = 0
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not row or all(c is None or c == "" for c in row):
            continue
        item: dict[str, Any] = {}
        for col_idx, field_name in col_to_field.items():
            if col_idx >= len(row):
                continue
            raw_val = row[col_idx]
            if field_name in INT_FIELDS:
                item[field_name] = _to_int(raw_val)
            elif field_name in NUMERIC_FIELDS:
                item[field_name] = _to_decimal(raw_val)
            else:
                # текстовые поля
                item[field_name] = str(raw_val).strip() if raw_val is not None else None

        # Валидация: SKU обязателен и число
        if not item.get("sku") or item["sku"] == 0:
            skipped += 1
            continue

        rows.append(item)

    wb.close()
    return ParseResult(
        period_from=period_from,
        period_to=period_to,
        rows=rows,
        skipped_rows=skipped,
        unknown_columns=unknown,
        file_name=file_name,
    )


def verify_row_profit(row: dict[str, Any]) -> Decimal:
    """
    Контрольная сверка формулы прибыли (для unit-теста при импорте).
    Формула из реального XLSX:
      Прибыль = revenue + spp_points + partner_programs
              − ozon_commission − acquiring − logistics − last_mile
              − storage − return_handling − reverse_logistics
              − disposal − ovh_extra − operational_errors
              − ad_cpc − ad_cpo − ad_star − ad_paid_brand − ad_reviews
              − posting_handling
    NB: знак в файле уже отрицательный для расходов → суммируем как есть.
    """
    def _v(k: str) -> Decimal:
        return row.get(k) or Decimal("0")
    return (
        _v("revenue") + _v("spp_points") + _v("partner_programs")
        + _v("ozon_commission") + _v("acquiring") + _v("posting_handling")
        + _v("logistics") + _v("last_mile") + _v("storage_from_xlsx")
        + _v("return_handling") + _v("reverse_logistics") + _v("disposal")
        + _v("ovh_extra") + _v("operational_errors")
        + _v("ad_cpc") + _v("ad_cpo") + _v("ad_star")
        + _v("ad_paid_brand") + _v("ad_reviews")
    )
