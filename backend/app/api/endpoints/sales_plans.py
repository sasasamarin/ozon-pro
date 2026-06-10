"""
/api/v1/plans — план продаж (новый раздел).

Endpoints:
  POST   /plans/forecast              — посчитать прогноз без сохранения
  POST   /plans/distribute            — распределить target по SKU
  POST   /plans                       — создать план (с items + daily)
  GET    /plans                       — список планов компании
  GET    /plans/{plan_id}             — детальный
  PATCH  /plans/{plan_id}             — обновить (метаданные / target)
  DELETE /plans/{plan_id}             — удалить
  PATCH  /plans/{plan_id}/items/{id}  — править item (plan_value/is_locked)
  POST   /plans/{plan_id}/rebalance   — пересчёт долей без lock-нутых
  POST   /plans/{plan_id}/distribute-days — рассчитать daily values
  POST   /plans/simulate              — симуляция каскада «+delta к метрике»
  GET    /plans/{plan_id}/fact        — факт + bridge + run-rate
  POST   /plans/{plan_id}/kpi         — назначить KPI менеджеру
  GET    /plans/{plan_id}/kpi         — список KPI плана
"""
from __future__ import annotations

import uuid
from datetime import date, datetime, timedelta
from decimal import Decimal
from typing import Any

import io

import openpyxl
from fastapi import APIRouter, Body, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import get_current_user
from app.api.deps_cabinets import get_accessible_cabinet_ids
from app.db.session import get_db
from app.models import User
from app.models.sales_plan import PlanKPI, SalesPlan, SalesPlanDaily, SalesPlanItem
from app.services.sales_plan.cascade import simulate_plan_change
from app.services.sales_plan.fact import compute_fact
from app.services.sales_plan.rbac import filter_cabinet_ids
from app.services.sales_plan.forecast import (
    compute_forecast, distribute_by_sku, distribute_by_sku_bottomup, fetch_history,
)


router = APIRouter()


# ============================================
# Schemas
# ============================================

class ForecastRequest(BaseModel):
    metric: str
    analysis_start: date
    analysis_end: date
    forecast_start: date
    forecast_end: date
    target_value: float | None = None
    cabinet_id: str | None = None
    product_id: str | None = None


class ForecastPoint(BaseModel):
    date: str
    value: float


class ForecastResponse(BaseModel):
    metric: str
    history: list[ForecastPoint]
    base_forecast: float
    forecast_series: list[ForecastPoint]
    modified_series: list[ForecastPoint] | None
    season_weights: dict[str, float]
    reliability: str
    reliability_pct: float
    note: str


class DistributeRequest(BaseModel):
    metric: str
    analysis_start: date
    analysis_end: date
    target_value: float
    cabinet_id: str | None = None


class BottomupRequest(BaseModel):
    """Bottom-up: выбираем кабинеты+товары, получаем стартовый прогноз per-SKU."""
    metric: str
    analysis_start: date
    analysis_end: date
    forecast_start: date
    forecast_end: date
    cabinet_ids: list[str] = []
    product_ids: list[str] = []


class DistributeItem(BaseModel):
    product_id: str | None
    sku: str | None
    name: str | None
    offer_id: str | None
    analysis_value: float
    share_pct: float
    plan_value: float


class BottomupItem(BaseModel):
    product_id: str | None
    sku: str | None
    name: str | None
    offer_id: str | None
    cabinet_id: str | None
    cabinet_name: str | None
    analysis_value: float
    analysis_value_clean: float | None = None
    outlier_excluded: float = 0
    normal_days: int = 0
    forecast_value: float
    plan_value: float
    share_pct: float


class BottomupResponse(BaseModel):
    items: list[BottomupItem]
    total_analysis: float
    total_forecast: float
    by_cabinet: list[dict]


class PlanCreate(BaseModel):
    name: str
    scope_type: str
    scope_ref: str | None = None
    metric_code: str
    period_start: date
    period_end: date
    analysis_start: date
    analysis_end: date
    target_value: float
    base_forecast: float | None = None
    distribution_mode: str = "proportional"
    source_pref: str = "operational"
    note: str | None = None
    items: list[DistributeItem] = []
    workspace_notes: str | None = None
    manual_adjustment: float = 0.0
    is_template: bool = False
    template_cabinet_ids: list[str] | None = None


class PlanItemRow(BaseModel):
    id: str
    product_id: str | None
    sku: str | None
    name: str | None
    offer_id: str | None
    analysis_value: float
    share_pct: float
    plan_value: float
    is_locked: bool


class PlanRow(BaseModel):
    id: str
    name: str
    scope_type: str
    scope_ref: str | None
    metric_code: str
    period_start: str
    period_end: str
    analysis_start: str
    analysis_end: str
    target_value: float
    base_forecast: float | None
    distribution_mode: str
    source_pref: str
    status: str
    note: str | None
    is_template: bool = False
    template_cabinet_ids: list[str] | None = None
    workspace_notes: str | None = None
    manual_adjustment: float = 0.0
    rolled_from_id: str | None = None
    created_at: str
    items_count: int


class PlanDetail(PlanRow):
    items: list[PlanItemRow]


class PlanUpdate(BaseModel):
    name: str | None = None
    target_value: float | None = None
    distribution_mode: str | None = None
    status: str | None = None  # draft | active | archived
    note: str | None = None
    workspace_notes: str | None = None
    manual_adjustment: float | None = None
    is_template: bool | None = None
    template_cabinet_ids: list[str] | None = None


class ItemUpdate(BaseModel):
    plan_value: float | None = None
    is_locked: bool | None = None


class SimulateRequest(BaseModel):
    metric: str
    delta: float
    period_start: date | None = None
    period_end: date | None = None
    cabinet_id: str | None = None


class KPICreate(BaseModel):
    manager_name: str
    metric_code: str
    target_value: float
    bonus_rule: dict | None = None


# ============================================
# Helpers
# ============================================

async def _get_plan_owned(
    db: AsyncSession, plan_id: uuid.UUID, company_id: uuid.UUID,
) -> SalesPlan:
    plan = (await db.execute(
        select(SalesPlan).where(
            SalesPlan.id == plan_id, SalesPlan.company_id == company_id,
        )
    )).scalar_one_or_none()
    if not plan:
        raise HTTPException(404, "План не найден")
    return plan


def _plan_to_row(plan: SalesPlan, items_count: int = 0) -> PlanRow:
    return PlanRow(
        id=str(plan.id), name=plan.name,
        scope_type=plan.scope_type, scope_ref=plan.scope_ref,
        metric_code=plan.metric_code,
        period_start=plan.period_start.isoformat(),
        period_end=plan.period_end.isoformat(),
        analysis_start=plan.analysis_start.isoformat(),
        analysis_end=plan.analysis_end.isoformat(),
        target_value=float(plan.target_value),
        base_forecast=float(plan.base_forecast) if plan.base_forecast else None,
        distribution_mode=plan.distribution_mode,
        source_pref=plan.source_pref,
        status=plan.status or "active",
        note=plan.note,
        is_template=bool(getattr(plan, "is_template", False)),
        template_cabinet_ids=(
            [str(c) for c in (plan.template_cabinet_ids or [])]
            if getattr(plan, "template_cabinet_ids", None) else None
        ),
        workspace_notes=getattr(plan, "workspace_notes", None),
        manual_adjustment=float(getattr(plan, "manual_adjustment", 0) or 0),
        rolled_from_id=str(getattr(plan, "rolled_from_id", None))
                       if getattr(plan, "rolled_from_id", None) else None,
        created_at=plan.created_at.isoformat(),
        items_count=items_count,
    )


def _item_to_row(item: SalesPlanItem, name: str | None = None) -> PlanItemRow:
    return PlanItemRow(
        id=str(item.id),
        product_id=str(item.product_id) if item.product_id else None,
        sku=item.sku,
        name=name,
        offer_id=item.sku,
        analysis_value=float(item.analysis_value),
        share_pct=float(item.share_pct),
        plan_value=float(item.plan_value),
        is_locked=item.is_locked,
    )


# ============================================
# Forecast / Distribute (без сохранения)
# ============================================

@router.post("/data-availability")
async def data_availability(
    payload: ForecastRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """🟢🟡🔴 по месяцам — где данные надёжные для прогноза.

    Возвращает по каждому месяцу analysis-периода:
      coverage_pct, status (green/yellow/red), days с данными, summa
    """
    from sqlalchemy import text as _sql
    from datetime import timedelta as _td

    cabinet = uuid.UUID(payload.cabinet_id) if payload.cabinet_id else None
    extra = "AND oa.id = :cab" if cabinet else ""
    params: dict = {"cid": str(current_user.company_id),
                    "df": payload.analysis_start, "dt": payload.analysis_end}
    if cabinet:
        params["cab"] = str(cabinet)

    if payload.metric == "revenue":
        sql = f"""
            SELECT t.operation_date::date AS day,
                   SUM(t.accruals_for_sale)::float AS v
            FROM transactions t
            JOIN ozon_accounts oa ON oa.id = t.ozon_account_id
            WHERE oa.company_id = :cid
              AND t.operation_date >= :df AND t.operation_date <= :dt
              AND t.operation_type='OperationAgentDeliveredToCustomer'
              {extra}
            GROUP BY 1 ORDER BY 1
        """
    elif payload.metric == "orders":
        sql = f"""
            SELECT o.order_created_at::date AS day, COUNT(*)::float AS v
            FROM orders o
            JOIN ozon_accounts oa ON oa.id = o.ozon_account_id
            WHERE oa.company_id = :cid AND o.status='delivered'
              AND o.order_created_at >= :df AND o.order_created_at <= :dt
              {extra}
            GROUP BY 1 ORDER BY 1
        """
    elif payload.metric == "units":
        sql = f"""
            SELECT o.order_created_at::date AS day, SUM(oi.quantity)::float AS v
            FROM order_items oi
            JOIN orders o ON o.id = oi.order_id
            JOIN ozon_accounts oa ON oa.id = o.ozon_account_id
            WHERE oa.company_id = :cid AND o.status='delivered'
              AND o.order_created_at >= :df AND o.order_created_at <= :dt
              {extra}
            GROUP BY 1 ORDER BY 1
        """
    else:
        sql = f"""
            SELECT t.operation_date::date AS day, COUNT(*)::float AS v
            FROM transactions t
            JOIN ozon_accounts oa ON oa.id = t.ozon_account_id
            WHERE oa.company_id = :cid
              AND t.operation_date >= :df AND t.operation_date <= :dt
              {extra}
            GROUP BY 1 ORDER BY 1
        """

    rows = (await db.execute(_sql(sql), params)).all()
    by_day = {r.day: float(r.v or 0) for r in rows}

    months: dict[str, dict] = {}
    d = payload.analysis_start
    while d <= payload.analysis_end:
        key = f"{d.year:04d}-{d.month:02d}"
        if key not in months:
            months[key] = {"days_with_data": 0, "days_total": 0, "sum_value": 0}
        months[key]["days_total"] += 1
        v = by_day.get(d, 0)
        if v > 0:
            months[key]["days_with_data"] += 1
            months[key]["sum_value"] += v
        d += _td(days=1)

    out = []
    for key, m in months.items():
        coverage = m["days_with_data"] / m["days_total"] * 100 if m["days_total"] else 0
        status = "green" if coverage >= 80 else ("yellow" if coverage >= 40 else "red")
        out.append({
            "month": key,
            "days_total": m["days_total"],
            "days_with_data": m["days_with_data"],
            "coverage_pct": round(coverage, 1),
            "status": status,
            "value": round(m["sum_value"], 2),
        })
    return {"months": out, "metric": payload.metric}


@router.post("/forecast", response_model=ForecastResponse)
async def forecast_plan(
    payload: ForecastRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> ForecastResponse:
    """Прогноз метрики на forecast-период по истории analysis-периода."""
    cabinet = uuid.UUID(payload.cabinet_id) if payload.cabinet_id else None
    product = uuid.UUID(payload.product_id) if payload.product_id else None
    history = await fetch_history(
        db, company_id=current_user.company_id, metric=payload.metric,
        analysis_start=payload.analysis_start, analysis_end=payload.analysis_end,
        cabinet_id=cabinet, product_id=product,
    )
    result = compute_forecast(
        history, payload.forecast_start, payload.forecast_end, payload.target_value,
    )
    return ForecastResponse(
        metric=payload.metric,
        history=[ForecastPoint(date=d.isoformat(), value=v) for d, v in result.history],
        base_forecast=result.base_forecast,
        forecast_series=[ForecastPoint(date=d.isoformat(), value=v) for d, v in result.forecast_series],
        modified_series=[ForecastPoint(date=d.isoformat(), value=v) for d, v in result.modified_series] if result.modified_series else None,
        season_weights=result.season_weights,
        reliability=result.reliability,
        reliability_pct=result.reliability_pct,
        note=result.note,
    )


@router.post("/bottomup", response_model=BottomupResponse)
async def bottomup_distribute(
    payload: BottomupRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> BottomupResponse:
    """Bottom-up: per-SKU прогноз → стартовые plan_value.

    Юзер потом правит вручную, авто-сумма обновляется на фронте.
    """
    # Защита от инвертированных периодов
    if payload.analysis_end < payload.analysis_start:
        raise HTTPException(400, "analysis_end раньше analysis_start")
    if payload.forecast_end < payload.forecast_start:
        raise HTTPException(400, "forecast_end раньше forecast_start")

    requested = [uuid.UUID(c) for c in payload.cabinet_ids] if payload.cabinet_ids else None
    cabs = await filter_cabinet_ids(db, current_user, requested)
    if cabs is not None and len(cabs) == 0:
        raise HTTPException(403, "Нет доступа к выбранным кабинетам")
    pids = [uuid.UUID(p) for p in payload.product_ids] if payload.product_ids else None
    items = await distribute_by_sku_bottomup(
        db, company_id=current_user.company_id, metric=payload.metric,
        analysis_start=payload.analysis_start, analysis_end=payload.analysis_end,
        forecast_start=payload.forecast_start, forecast_end=payload.forecast_end,
        cabinet_ids=cabs, product_ids=pids,
    )

    # Группировка по кабинету
    by_cab: dict[str, dict] = {}
    for it in items:
        cid = it["cabinet_id"] or "—"
        if cid not in by_cab:
            by_cab[cid] = {
                "cabinet_id": it["cabinet_id"],
                "cabinet_name": it["cabinet_name"] or "(без кабинета)",
                "analysis_sum": 0, "forecast_sum": 0, "plan_sum": 0,
                "skus_count": 0,
            }
        by_cab[cid]["analysis_sum"] += it["analysis_value"]
        by_cab[cid]["forecast_sum"] += it["forecast_value"]
        by_cab[cid]["plan_sum"] += it["plan_value"]
        by_cab[cid]["skus_count"] += 1
    by_cabinet = [
        {**v, "analysis_sum": round(v["analysis_sum"], 2),
         "forecast_sum": round(v["forecast_sum"], 2),
         "plan_sum": round(v["plan_sum"], 2)}
        for v in by_cab.values()
    ]

    if len(items) == 0:
        # Подсказка юзеру: товаров с продажами не нашли. Возможные причины.
        from sqlalchemy import text as _sql
        chk_oa = (await db.execute(_sql(
            "SELECT COUNT(*) FROM ozon_accounts WHERE company_id = :cid AND deleted_at IS NULL"
        ), {"cid": str(current_user.company_id)})).scalar()
        chk_orders = (await db.execute(_sql("""
            SELECT COUNT(*) FROM orders o JOIN ozon_accounts oa ON oa.id = o.ozon_account_id
            WHERE oa.company_id = :cid AND o.order_created_at >= :df AND o.order_created_at <= :dt
        """), {"cid": str(current_user.company_id),
               "df": payload.analysis_start, "dt": payload.analysis_end})).scalar()
        chk_tx = (await db.execute(_sql("""
            SELECT COUNT(*) FROM transactions t JOIN ozon_accounts oa ON oa.id = t.ozon_account_id
            WHERE oa.company_id = :cid AND t.operation_date >= :df AND t.operation_date <= :dt
        """), {"cid": str(current_user.company_id),
               "df": payload.analysis_start, "dt": payload.analysis_end})).scalar()
        raise HTTPException(404, detail={
            "message": "Нет данных по выбранным фильтрам",
            "cabinets_in_company": chk_oa,
            "cabinets_selected": len(cabs) if cabs else "all",
            "orders_in_period": chk_orders,
            "transactions_in_period": chk_tx,
            "hint": (
                "Проверь: 1) подключены ли кабинеты с данными; "
                "2) сделана ли синхронизация за выбранный период; "
                "3) есть ли у роли доступ к этим кабинетам (MemberAccountAccess)."
            ),
        })

    return BottomupResponse(
        items=[BottomupItem(**i) for i in items],
        total_analysis=round(sum(i["analysis_value"] for i in items), 2),
        total_forecast=round(sum(i["forecast_value"] for i in items), 2),
        by_cabinet=by_cabinet,
    )


@router.post("/distribute", response_model=list[DistributeItem])
async def distribute_plan(
    payload: DistributeRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> list[DistributeItem]:
    """Распределить target по SKU пропорционально вкладу в analysis-периоде."""
    cabinet = uuid.UUID(payload.cabinet_id) if payload.cabinet_id else None
    items = await distribute_by_sku(
        db, company_id=current_user.company_id, metric=payload.metric,
        analysis_start=payload.analysis_start, analysis_end=payload.analysis_end,
        cabinet_id=cabinet, target_value=payload.target_value,
    )
    return [DistributeItem(**i) for i in items]


# ============================================
# CRUD планов
# ============================================

@router.post("", response_model=PlanDetail)
async def create_plan(
    payload: PlanCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> PlanDetail:
    """Создать план + сохранить items + (опционально) рассчитать daily."""
    plan = SalesPlan(
        company_id=current_user.company_id, user_id=current_user.id,
        name=payload.name, scope_type=payload.scope_type,
        scope_ref=payload.scope_ref, metric_code=payload.metric_code,
        period_start=payload.period_start, period_end=payload.period_end,
        analysis_start=payload.analysis_start, analysis_end=payload.analysis_end,
        target_value=Decimal(str(payload.target_value)),
        base_forecast=(Decimal(str(payload.base_forecast))
                       if payload.base_forecast is not None else None),
        distribution_mode=payload.distribution_mode,
        source_pref=payload.source_pref,
        note=payload.note,
        is_template=payload.is_template,
        template_cabinet_ids=payload.template_cabinet_ids,
        workspace_notes=payload.workspace_notes,
        manual_adjustment=Decimal(str(payload.manual_adjustment or 0)),
    )
    db.add(plan)
    await db.flush()

    for it in payload.items:
        db.add(SalesPlanItem(
            plan_id=plan.id,
            product_id=uuid.UUID(it.product_id) if it.product_id else None,
            sku=it.sku,
            analysis_value=Decimal(str(it.analysis_value)),
            share_pct=Decimal(str(it.share_pct)),
            plan_value=Decimal(str(it.plan_value)),
        ))
    await db.commit()
    await db.refresh(plan)

    items = (await db.execute(
        select(SalesPlanItem).where(SalesPlanItem.plan_id == plan.id)
    )).scalars().all()
    return PlanDetail(
        **_plan_to_row(plan, len(items)).model_dump(),
        items=[_item_to_row(i, name=None) for i in items],
    )


@router.get("", response_model=list[PlanRow])
async def list_plans(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> list[PlanRow]:
    plans = (await db.execute(
        select(SalesPlan).where(SalesPlan.company_id == current_user.company_id)
        .order_by(SalesPlan.period_start.desc())
    )).scalars().all()

    # Карта кабинетов для подстановки названий в template_cabinet_ids
    from app.models import OzonAccount
    accessible = await get_accessible_cabinet_ids(db, current_user)
    cabs_q = select(OzonAccount.id, OzonAccount.name).where(
        OzonAccount.company_id == current_user.company_id,
    )
    if accessible is not None:
        cabs_q = cabs_q.where(OzonAccount.id.in_(accessible))
    cabs_rows = (await db.execute(cabs_q)).all()
    cab_name = {str(r.id): r.name for r in cabs_rows}

    rows = []
    for p in plans:
        cnt = (await db.execute(
            select(SalesPlanItem).where(SalesPlanItem.plan_id == p.id)
        )).scalars().all()
        row = _plan_to_row(p, len(cnt))
        # Шаблон → подставить имена кабинетов в template_cabinet_ids
        if row.template_cabinet_ids:
            row.template_cabinet_ids = [
                cab_name.get(cid, cid) for cid in row.template_cabinet_ids
            ]
        rows.append(row)
    return rows


@router.get("/{plan_id}", response_model=PlanDetail)
async def get_plan(
    plan_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> PlanDetail:
    try:
        pid = uuid.UUID(plan_id)
    except ValueError:
        raise HTTPException(400, "Невалидный id")
    plan = await _get_plan_owned(db, pid, current_user.company_id)
    items = (await db.execute(
        select(SalesPlanItem).where(SalesPlanItem.plan_id == plan.id)
    )).scalars().all()
    # Имена товаров
    name_map = {}
    pids = [i.product_id for i in items if i.product_id]
    if pids:
        from app.models import Product
        rows = (await db.execute(
            select(Product.id, Product.name).where(Product.id.in_(pids))
        )).all()
        name_map = {r.id: r.name for r in rows}
    return PlanDetail(
        **_plan_to_row(plan, len(items)).model_dump(),
        items=[_item_to_row(i, name=name_map.get(i.product_id)) for i in items],
    )


@router.patch("/{plan_id}", response_model=PlanRow)
async def update_plan(
    plan_id: str,
    payload: PlanUpdate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> PlanRow:
    try:
        pid = uuid.UUID(plan_id)
    except ValueError:
        raise HTTPException(400, "Невалидный id")
    plan = await _get_plan_owned(db, pid, current_user.company_id)
    if payload.name is not None:
        plan.name = payload.name
    if payload.target_value is not None:
        plan.target_value = Decimal(str(payload.target_value))
    if payload.distribution_mode is not None:
        plan.distribution_mode = payload.distribution_mode
    if payload.status is not None:
        plan.status = payload.status
    if payload.note is not None:
        plan.note = payload.note
    if payload.workspace_notes is not None:
        plan.workspace_notes = payload.workspace_notes
    if payload.manual_adjustment is not None:
        plan.manual_adjustment = Decimal(str(payload.manual_adjustment))
    if payload.is_template is not None:
        plan.is_template = payload.is_template
    if payload.template_cabinet_ids is not None:
        plan.template_cabinet_ids = payload.template_cabinet_ids
    plan.updated_at = datetime.now(tz=plan.updated_at.tzinfo if plan.updated_at.tzinfo else None)
    await db.commit()
    return _plan_to_row(plan)


@router.delete("/{plan_id}")
async def delete_plan(
    plan_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    try:
        pid = uuid.UUID(plan_id)
    except ValueError:
        raise HTTPException(400, "Невалидный id")
    plan = await _get_plan_owned(db, pid, current_user.company_id)
    await db.delete(plan)
    await db.commit()
    return {"ok": True}


@router.patch("/{plan_id}/items/{item_id}", response_model=PlanItemRow)
async def update_item(
    plan_id: str,
    item_id: str,
    payload: ItemUpdate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> PlanItemRow:
    try:
        pid = uuid.UUID(plan_id)
        iid = uuid.UUID(item_id)
    except ValueError:
        raise HTTPException(400, "Невалидный id")
    await _get_plan_owned(db, pid, current_user.company_id)
    item = (await db.execute(
        select(SalesPlanItem).where(
            SalesPlanItem.id == iid, SalesPlanItem.plan_id == pid,
        )
    )).scalar_one_or_none()
    if not item:
        raise HTTPException(404, "Позиция не найдена")
    if payload.plan_value is not None:
        item.plan_value = Decimal(str(payload.plan_value))
    if payload.is_locked is not None:
        item.is_locked = payload.is_locked
    await db.commit()
    return _item_to_row(item)


@router.post("/{plan_id}/rebalance", response_model=PlanDetail)
async def rebalance(
    plan_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> PlanDetail:
    """Перераспределить НЕ-залоченные items так, чтобы сумма = target."""
    try:
        pid = uuid.UUID(plan_id)
    except ValueError:
        raise HTTPException(400, "Невалидный id")
    plan = await _get_plan_owned(db, pid, current_user.company_id)
    items = (await db.execute(
        select(SalesPlanItem).where(SalesPlanItem.plan_id == pid)
    )).scalars().all()

    locked_sum = sum(float(i.plan_value) for i in items if i.is_locked)
    target = float(plan.target_value)
    remaining = max(0.0, target - locked_sum)
    unlocked_analysis = sum(float(i.analysis_value) for i in items if not i.is_locked)

    if unlocked_analysis > 0:
        for i in items:
            if i.is_locked:
                continue
            share = float(i.analysis_value) / unlocked_analysis
            i.plan_value = Decimal(str(round(remaining * share, 2)))
            i.share_pct = Decimal(str(round(share * 100, 4)))
    await db.commit()
    return await get_plan(plan_id=str(pid), current_user=current_user, db=db)


@router.get("/{plan_id}/weeks")
async def items_by_weeks(
    plan_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Сетка SKU × недели для шага 3 визарда.

    Возвращает:
      weeks: список ISO-недель [{week_start, week_end, label}]
      rows: [{item_id, sku, name, plan_value, weeks: [{week_start, value}]}]
    """
    from datetime import timedelta as _td
    try:
        pid = uuid.UUID(plan_id)
    except ValueError:
        raise HTTPException(400, "Невалидный id")
    plan = await _get_plan_owned(db, pid, current_user.company_id)

    # Список недель в периоде
    weeks = []
    cursor = plan.period_start - _td(days=plan.period_start.weekday())  # Mon
    while cursor <= plan.period_end:
        wend = min(cursor + _td(days=6), plan.period_end)
        weeks.append({
            "week_start": cursor.isoformat(),
            "week_end": wend.isoformat(),
            "label": f"{cursor.day:02d}.{cursor.month:02d}",
        })
        cursor += _td(days=7)

    # Items с агрегацией daily по неделям
    items = (await db.execute(
        select(SalesPlanItem).where(SalesPlanItem.plan_id == pid)
        .order_by(SalesPlanItem.plan_value.desc())
    )).scalars().all()

    from app.models import Product
    pids = [i.product_id for i in items if i.product_id]
    name_map = {}
    if pids:
        rows = (await db.execute(
            select(Product.id, Product.name).where(Product.id.in_(pids))
        )).all()
        name_map = {r.id: r.name for r in rows}

    rows_out = []
    for it in items:
        daily = (await db.execute(
            select(SalesPlanDaily).where(SalesPlanDaily.plan_item_id == it.id)
        )).scalars().all()
        # Группировка по неделям
        weeks_data = []
        for w in weeks:
            ws = date.fromisoformat(w["week_start"])
            we = date.fromisoformat(w["week_end"])
            total = sum(
                float(d.plan_value) for d in daily
                if ws <= d.date <= we
            )
            weeks_data.append({"week_start": w["week_start"], "value": round(total, 2)})
        rows_out.append({
            "item_id": str(it.id),
            "product_id": str(it.product_id) if it.product_id else None,
            "sku": it.sku, "name": name_map.get(it.product_id) or it.sku,
            "plan_value": float(it.plan_value),
            "weeks": weeks_data,
        })
    return {"weeks": weeks, "rows": rows_out}


@router.get("/{plan_id}/stock-hint")
async def stock_hint(
    plan_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Мягкий hint о складе. НЕ блокирует план, только показывает дефицит.

    Доступно = остаток + в пути + плановые поставки (со сроками).
    Если план_шт > доступно → флаг hint 🟡 «нужен товар ~N шт».
    """
    try:
        pid = uuid.UUID(plan_id)
    except ValueError:
        raise HTTPException(400, "Невалидный id")
    plan = await _get_plan_owned(db, pid, current_user.company_id)

    items = (await db.execute(
        select(SalesPlanItem).where(SalesPlanItem.plan_id == pid)
    )).scalars().all()

    hints = []
    from sqlalchemy import text as _sql
    for it in items:
        if not it.product_id:
            continue
        plan_units = float(it.plan_value)
        # Для метрики revenue/gross_profit — конверсия plan_units из ₽ в шт.
        # Если метрика orders/units — plan_value уже в штуках.
        # Простой случай: для revenue делим на seller_price.
        if plan.metric_code in ("revenue", "gross_profit"):
            sp_row = (await db.execute(_sql(
                "SELECT marketing_seller_price FROM products WHERE id = :pid"
            ), {"pid": str(it.product_id)})).first()
            sp = float(sp_row.marketing_seller_price or 0) if sp_row else 0
            if sp > 0:
                plan_units = plan_units / sp

        stock_row = (await db.execute(_sql("""
            SELECT COALESCE(SUM(stock_for_sale), 0)::float AS stock
            FROM warehouse_stocks
            WHERE product_id = :pid
              AND snapshot_date = (SELECT MAX(snapshot_date) FROM warehouse_stocks)
        """), {"pid": str(it.product_id)})).first()
        stock = float(stock_row.stock or 0) if stock_row else 0

        # Плановые поставки (supplies) до конца плана-периода
        supply_row = (await db.execute(_sql("""
            SELECT COALESCE(SUM(si.qty), 0)::int AS qty,
                   MIN(s.expected_date)::date AS earliest
            FROM supply_items si
            JOIN supplies s ON s.id = si.supply_id
            WHERE si.product_id = :pid
              AND (s.expected_date IS NULL OR s.expected_date <= :end)
        """), {"pid": str(it.product_id), "end": plan.period_end})).first()
        in_supply = int(supply_row.qty or 0) if supply_row else 0
        earliest = supply_row.earliest if supply_row else None

        available = stock + in_supply
        if plan_units > available:
            deficit = round(plan_units - available)
            hints.append({
                "product_id": str(it.product_id),
                "sku": it.sku,
                "plan_units": round(plan_units),
                "stock_now": round(stock),
                "in_supply": in_supply,
                "available": round(available),
                "deficit_units": deficit,
                "earliest_supply": earliest.isoformat() if earliest else None,
                "message": (
                    f"🟡 Нужен товар ~{deficit} шт"
                    + (f", поставка к {earliest.isoformat()}" if earliest
                       else "; нет плановых поставок")
                ),
            })

    return {
        "plan_id": str(pid),
        "hints": hints,
        "blocked": False,  # mainly informational
        "note": "Подсказка не блокирует план. Закупки — отдельный раздел.",
    }


@router.post("/{plan_id}/distribute-days")
async def distribute_days(
    plan_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Раскинуть plan_value каждого item по дням периода с сезонными весами."""
    try:
        pid = uuid.UUID(plan_id)
    except ValueError:
        raise HTTPException(400, "Невалидный id")
    plan = await _get_plan_owned(db, pid, current_user.company_id)

    history = await fetch_history(
        db, company_id=current_user.company_id, metric=plan.metric_code,
        analysis_start=plan.analysis_start, analysis_end=plan.analysis_end,
    )
    fc = compute_forecast(
        history, plan.period_start, plan.period_end,
        target_value=float(plan.target_value),
    )

    items = (await db.execute(
        select(SalesPlanItem).where(SalesPlanItem.plan_id == pid)
    )).scalars().all()

    # Удаляем старые daily
    for it in items:
        await db.execute(
            delete(SalesPlanDaily).where(SalesPlanDaily.plan_item_id == it.id)
        )
    await db.flush()

    total_days = 0
    for it in items:
        for d_iso, weight in fc.season_weights.items():
            db.add(SalesPlanDaily(
                plan_item_id=it.id,
                date=date.fromisoformat(d_iso),
                plan_value=Decimal(str(round(float(it.plan_value) * weight, 2))),
                season_weight=Decimal(str(round(weight, 6))),
            ))
            total_days += 1
    await db.commit()
    return {"daily_rows": total_days, "reliability": fc.reliability}


# ============================================
# Simulate cascade
# ============================================

@router.post("/simulate")
async def simulate(
    payload: SimulateRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """«+10000 кликов» → каскад с эффектами на orders/revenue/margin/ДРР."""
    cabinet = uuid.UUID(payload.cabinet_id) if payload.cabinet_id else None
    result = await simulate_plan_change(
        db, company_id=current_user.company_id, metric=payload.metric,
        delta=payload.delta, period_start=payload.period_start,
        period_end=payload.period_end, cabinet_id=cabinet,
    )
    return {
        "input_metric": result.input_metric,
        "input_delta": result.input_delta,
        "base_period": {
            "from": result.base_period[0].isoformat(),
            "to": result.base_period[1].isoformat(),
        },
        "effects": [
            {"metric": e.metric, "delta": e.delta,
             "new_value": e.new_value, "explanation": e.explanation}
            for e in result.effects
        ],
        "cpc_breakeven": result.cpc_breakeven,
        "drr_before_pct": result.drr_before_pct,
        "drr_after_pct": result.drr_after_pct,
        "note": result.note,
    }


# ============================================
# Fact + bridge
# ============================================

@router.get("/{plan_id}/fact")
async def get_fact(
    plan_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    try:
        pid = uuid.UUID(plan_id)
    except ValueError:
        raise HTTPException(400, "Невалидный id")
    plan = await _get_plan_owned(db, pid, current_user.company_id)
    cabinet = uuid.UUID(plan.scope_ref) if plan.scope_type == "cabinet" and plan.scope_ref else None
    result = await compute_fact(
        db, company_id=current_user.company_id,
        plan_value=float(plan.target_value), metric=plan.metric_code,
        period_start=plan.period_start, period_end=plan.period_end,
        cabinet_id=cabinet,
    )
    return {
        "plan_value": result.plan_value,
        "fact_value": result.fact_value,
        "fact_source": result.fact_source,
        "is_preliminary": result.is_preliminary,
        "delta_realization_tx": result.delta_realization_tx,
        "completion_pct": result.completion_pct,
        "completion_prorata_pct": result.completion_prorata_pct,
        "run_rate_forecast": result.run_rate_forecast,
        "needed_per_day": result.needed_per_day,
        "days_elapsed": result.days_elapsed,
        "days_remaining": result.days_remaining,
        "days_total": result.days_total,
        "probability_pct": result.probability_pct,
        "bridge": [
            {"name": b.name, "value": b.value, "explanation": b.explanation}
            for b in result.bridge
        ],
        "note": result.note,
        # Структура отчёта Ozon: Оплачено(брутто) − Возвращено = Выручка(нетто)
        "revenue_breakdown": {
            "gross": result.gross_revenue,
            "returns": result.returns_amount,
            "returns_count": result.returns_count,
            "net": result.net_revenue,
            "formula": "Оплачено (брутто) − Возвращено = Выручка (нетто)",
        },
    }


# ============================================
# KPI
# ============================================

@router.get("/{plan_id}/items.xlsx")
async def export_items_xlsx(
    plan_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Excel-экспорт распределения по SKU. Можно править в Excel и заливать обратно."""
    try:
        pid = uuid.UUID(plan_id)
    except ValueError:
        raise HTTPException(400, "Невалидный id")
    plan = await _get_plan_owned(db, pid, current_user.company_id)
    items = (await db.execute(
        select(SalesPlanItem).where(SalesPlanItem.plan_id == pid)
        .order_by(SalesPlanItem.plan_value.desc())
    )).scalars().all()

    # Карта product_id → product_name
    from app.models import Product
    pids = [i.product_id for i in items if i.product_id]
    name_map: dict = {}
    if pids:
        rows = (await db.execute(
            select(Product.id, Product.name).where(Product.id.in_(pids))
        )).all()
        name_map = {r.id: r.name for r in rows}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "План"
    ws.append([
        "item_id", "product_id", "offer_id", "Товар",
        "Период анализа", "Доля %", "План значение", "Lock",
    ])
    for col, w in zip("ABCDEFGH", [38, 38, 18, 50, 16, 10, 16, 6]):
        ws.column_dimensions[col].width = w
    for it in items:
        ws.append([
            str(it.id), str(it.product_id) if it.product_id else "",
            it.sku or "", name_map.get(it.product_id, "")[:80],
            float(it.analysis_value), float(it.share_pct),
            float(it.plan_value), "lock" if it.is_locked else "",
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="plan_{plan.name[:30]}.xlsx"',
        },
    )


@router.post("/{plan_id}/items/import")
async def import_items_xlsx(
    plan_id: str,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Загрузка Excel: обновляет plan_value/is_locked по item_id.

    Только колонки G (План значение) и H (Lock) применяются. item_id строго
    проверяется на принадлежность плану. Возвращает {updated, skipped, errors}.
    """
    try:
        pid = uuid.UUID(plan_id)
    except ValueError:
        raise HTTPException(400, "Невалидный id")
    plan = await _get_plan_owned(db, pid, current_user.company_id)

    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(400, "Ожидается XLSX")
    content = await file.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(413, "Файл больше 5 МБ")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Не удалось открыть XLSX: {e}")
    ws = wb.active

    updated = 0
    skipped = 0
    errors: list[str] = []

    # Загружаем все items плана сразу
    items = (await db.execute(
        select(SalesPlanItem).where(SalesPlanItem.plan_id == pid)
    )).scalars().all()
    by_id = {str(it.id): it for it in items}

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue
        item_id = str(row[0]).strip()
        if item_id not in by_id:
            errors.append(f"row {row_idx}: item_id {item_id[:8]}… не найден")
            skipped += 1
            continue
        it = by_id[item_id]
        try:
            if row[6] is not None:
                it.plan_value = Decimal(str(row[6]))
            if row[7] is not None:
                it.is_locked = str(row[7]).lower() in ("lock", "true", "1", "да")
            updated += 1
        except Exception as e:
            errors.append(f"row {row_idx}: {e}")
            skipped += 1

    await db.commit()
    return {"updated": updated, "skipped": skipped, "errors": errors[:20]}


@router.post("/{plan_id}/clone", response_model=PlanRow)
async def clone_plan(
    plan_id: str,
    name: str | None = None,
    period_start: date | None = None,
    period_end: date | None = None,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Клонировать план как шаблон. Сохраняет items с теми же plan_value
    (можно потом перераспределить под новый период).

    Если задан новый период — items копируются, distribute-days
    запускается заново на frontend через POST /{new_id}/distribute-days.
    """
    try:
        src_id = uuid.UUID(plan_id)
    except ValueError:
        raise HTTPException(400, "Невалидный id")
    src = await _get_plan_owned(db, src_id, current_user.company_id)

    new = SalesPlan(
        company_id=current_user.company_id, user_id=current_user.id,
        name=name or f"{src.name} (копия)",
        scope_type=src.scope_type, scope_ref=src.scope_ref,
        metric_code=src.metric_code,
        period_start=period_start or src.period_start,
        period_end=period_end or src.period_end,
        analysis_start=src.analysis_start, analysis_end=src.analysis_end,
        target_value=src.target_value, base_forecast=src.base_forecast,
        distribution_mode=src.distribution_mode,
        source_pref=src.source_pref,
        status="draft",  # клон стартует как draft, юзер активирует
        note=f"Клонировано из {src.id}",
    )
    db.add(new)
    await db.flush()

    # Копируем items
    src_items = (await db.execute(
        select(SalesPlanItem).where(SalesPlanItem.plan_id == src_id)
    )).scalars().all()
    for it in src_items:
        db.add(SalesPlanItem(
            plan_id=new.id, product_id=it.product_id, sku=it.sku,
            analysis_value=it.analysis_value, share_pct=it.share_pct,
            plan_value=it.plan_value,
        ))
    await db.commit()
    await db.refresh(new)
    return _plan_to_row(new, len(src_items))


class BulkAction(BaseModel):
    plan_ids: list[str]
    action: str  # 'delete' | 'archive' | 'activate' | 'draft'


@router.post("/bulk")
async def bulk_action(
    payload: BulkAction,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Массовое действие над выбранными планами компании."""
    if payload.action not in ("delete", "archive", "activate", "draft"):
        raise HTTPException(400, f"Неизвестное действие: {payload.action}")
    try:
        pids = [uuid.UUID(p) for p in payload.plan_ids]
    except ValueError:
        raise HTTPException(400, "Невалидный plan_id в списке")

    if not pids:
        return {"affected": 0}

    plans = (await db.execute(
        select(SalesPlan).where(
            SalesPlan.id.in_(pids),
            SalesPlan.company_id == current_user.company_id,
        )
    )).scalars().all()

    affected = 0
    for p in plans:
        if payload.action == "delete":
            await db.delete(p)
        elif payload.action == "archive":
            p.status = "archived"
        elif payload.action == "activate":
            p.status = "active"
        elif payload.action == "draft":
            p.status = "draft"
        affected += 1
    await db.commit()
    return {"affected": affected, "action": payload.action}


@router.post("/rollover")
async def rollover_plans(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Idempotent rollover: закрывает все активные планы где period_end < today
    в архив. Если у плана есть rolled_from_id указывающий на шаблон —
    создаёт новый активный план в следующем периоде с теми же items.
    """
    from datetime import timedelta as _td
    today = date.today()
    closed = (await db.execute(
        select(SalesPlan).where(
            SalesPlan.company_id == current_user.company_id,
            SalesPlan.status == "active",
            SalesPlan.period_end < today,
            SalesPlan.is_template == False,
        )
    )).scalars().all()

    archived = 0
    rolled = 0
    new_plans: list[str] = []

    for plan in closed:
        plan.status = "archived"
        archived += 1
        if plan.rolled_from_id:
            tpl = (await db.execute(
                select(SalesPlan).where(SalesPlan.id == plan.rolled_from_id)
            )).scalar_one_or_none()
            if tpl and tpl.is_template:
                period_len = (plan.period_end - plan.period_start).days + 1
                new_start = plan.period_end + _td(days=1)
                new_end = new_start + _td(days=period_len - 1)
                new_plan = SalesPlan(
                    company_id=current_user.company_id, user_id=current_user.id,
                    name=f"{tpl.name} — {new_start.isoformat()}",
                    scope_type=tpl.scope_type, scope_ref=tpl.scope_ref,
                    metric_code=tpl.metric_code,
                    period_start=new_start, period_end=new_end,
                    analysis_start=plan.analysis_start,
                    analysis_end=plan.analysis_end,
                    target_value=tpl.target_value,
                    distribution_mode=tpl.distribution_mode,
                    source_pref=tpl.source_pref,
                    status="active",
                    rolled_from_id=tpl.id,
                )
                db.add(new_plan)
                await db.flush()
                src_items = (await db.execute(
                    select(SalesPlanItem).where(SalesPlanItem.plan_id == plan.id)
                )).scalars().all()
                for it in src_items:
                    db.add(SalesPlanItem(
                        plan_id=new_plan.id,
                        product_id=it.product_id, sku=it.sku,
                        analysis_value=it.analysis_value,
                        share_pct=it.share_pct, plan_value=it.plan_value,
                    ))
                rolled += 1
                new_plans.append(str(new_plan.id))
    await db.commit()
    return {"archived": archived, "rolled": rolled, "new_plan_ids": new_plans}


@router.get("/{plan_id}/overview")
async def fact_overview(
    plan_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Расширенный обзор факта: 3 колонки (Цель / Сейчас / Прогноз)
    × разрез по SKU с per-SKU performance.

    Возвращает:
      summary: {target, current, forecast_end_period, pct_current, pct_forecast}
      sku_rows: [{sku, name, target, current, forecast, pct_current, pct_forecast, tone}]
    """
    from datetime import timedelta as _td
    from sqlalchemy import text as _sql
    try:
        pid = uuid.UUID(plan_id)
    except ValueError:
        raise HTTPException(400, "Невалидный id")
    plan = await _get_plan_owned(db, pid, current_user.company_id)

    today = date.today()
    days_total = (plan.period_end - plan.period_start).days + 1
    days_elapsed = max(0, min(days_total, (today - plan.period_start).days + 1))

    cabinet = (uuid.UUID(plan.scope_ref)
               if plan.scope_type == "cabinet" and plan.scope_ref else None)
    extra = "AND oa.id = :cab" if cabinet else ""
    params = {"cid": str(current_user.company_id),
              "df": plan.period_start, "dt": min(today, plan.period_end)}
    if cabinet:
        params["cab"] = str(cabinet)

    # Получаем per-product factual из исторических данных
    # Дополним per-SKU plan
    items = (await db.execute(
        select(SalesPlanItem).where(SalesPlanItem.plan_id == pid)
    )).scalars().all()
    items_by_pid = {str(it.product_id): it for it in items if it.product_id}

    # SKU-факт за период плана. transactions нет product_id → через order_items.
    if plan.metric_code == "revenue":
        sql = f"""
            SELECT oi.product_id::text AS pid,
                   COALESCE(SUM(oi.price * oi.quantity), 0)::float AS v
            FROM order_items oi
            JOIN orders o ON o.id = oi.order_id
            JOIN ozon_accounts oa ON oa.id = o.ozon_account_id
            WHERE oa.company_id = :cid AND o.status='delivered'
              AND o.order_created_at >= :df AND o.order_created_at <= :dt
              {extra}
            GROUP BY oi.product_id
        """
    elif plan.metric_code in ("orders", "units"):
        agg = "COUNT(DISTINCT o.id)::float" if plan.metric_code == "orders" else "SUM(oi.quantity)::float"
        sql = f"""
            SELECT oi.product_id::text AS pid, {agg} AS v
            FROM order_items oi
            JOIN orders o ON o.id = oi.order_id
            JOIN ozon_accounts oa ON oa.id = o.ozon_account_id
            WHERE oa.company_id = :cid AND o.status='delivered'
              AND o.order_created_at >= :df AND o.order_created_at <= :dt
              {extra}
            GROUP BY oi.product_id
        """
    else:
        sql = f"""
            SELECT oi.product_id::text AS pid,
                   COALESCE(SUM(oi.price * oi.quantity), 0)::float AS v
            FROM order_items oi
            JOIN orders o ON o.id = oi.order_id
            JOIN ozon_accounts oa ON oa.id = o.ozon_account_id
            WHERE oa.company_id = :cid AND o.status='delivered'
              AND o.order_created_at >= :df AND o.order_created_at <= :dt
              {extra}
            GROUP BY oi.product_id
        """
    fact_rows = (await db.execute(_sql(sql), params)).all()
    fact_by_pid = {r.pid: float(r.v or 0) for r in fact_rows}

    # Имена SKU
    from app.models import Product
    all_pids = list(items_by_pid.keys())
    name_map = {}
    if all_pids:
        rows_p = (await db.execute(
            select(Product.id, Product.name, Product.offer_id).where(
                Product.id.in_([uuid.UUID(p) for p in all_pids])
            )
        )).all()
        name_map = {str(r.id): (r.name, r.offer_id) for r in rows_p}

    sku_rows = []
    total_target = 0.0
    total_current = 0.0
    total_forecast = 0.0
    prorata_factor = (days_elapsed / days_total) if days_total > 0 else 0
    for pid_str, it in items_by_pid.items():
        target = float(it.plan_value)
        current = fact_by_pid.get(pid_str, 0)
        forecast = (current / days_elapsed * days_total) if days_elapsed > 0 else 0
        pct_cur = (current / target * 100) if target > 0 else None
        pct_fc = (forecast / target * 100) if target > 0 else None
        prorata_target = target * prorata_factor
        deviation = current - prorata_target

        n, off = name_map.get(pid_str, ("", ""))
        tone = "rose"
        if pct_fc is not None:
            if pct_fc >= 100: tone = "emerald"
            elif pct_fc >= 80: tone = "amber"

        sku_rows.append({
            "product_id": pid_str,
            "sku": off or it.sku or "",
            "name": n or "(неизвестно)",
            "target": round(target, 2),
            "current": round(current, 2),
            "forecast": round(forecast, 2),
            "prorata_target": round(prorata_target, 2),
            "deviation": round(deviation, 2),
            "pct_current": round(pct_cur, 1) if pct_cur is not None else None,
            "pct_forecast": round(pct_fc, 1) if pct_fc is not None else None,
            "tone": tone,
        })
        total_target += target
        total_current += current
        total_forecast += forecast

    # Сортировка по |отклонению| — проблемные SKU наверху
    sku_rows.sort(key=lambda r: abs(r["deviation"] or 0), reverse=True)

    summary_pct_cur = (total_current / total_target * 100) if total_target > 0 else None
    summary_pct_fc = (total_forecast / total_target * 100) if total_target > 0 else None

    return {
        "plan_id": str(pid),
        "plan_name": plan.name,
        "metric_code": plan.metric_code,
        "period_start": plan.period_start.isoformat(),
        "period_end": plan.period_end.isoformat(),
        "days_elapsed": days_elapsed,
        "days_total": days_total,
        "summary": {
            "target": round(total_target, 2),
            "current": round(total_current, 2),
            "forecast": round(total_forecast, 2),
            "pct_current": round(summary_pct_cur, 1) if summary_pct_cur is not None else None,
            "pct_forecast": round(summary_pct_fc, 1) if summary_pct_fc is not None else None,
        },
        "sku_rows": sku_rows,
    }


@router.get("/{plan_id}/dashboard")
async def fact_dashboard(
    plan_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Сводный дашборд факта в 3 срезах × 3 группы метрик.

    Срезы: today (день) | cumulative (накопительно с начала плана) | month (текущий месяц)
    Группы:
      ПРОДАЖИ: заказы, выкуп, выручка-нетто
      ФИНАНСЫ: маржинальная прибыль, чистая прибыль (УСН 6% эвристика)
      ПРОДВИЖЕНИЕ: ДРР, рекламный расход
    Для каждой ячейки: {fact, plan_prorata, pct}.
    Светофор по pct: ≥100 emerald, ≥80 amber, <80 rose.
    """
    from datetime import timedelta as _td
    from sqlalchemy import text as _sql
    try:
        pid = uuid.UUID(plan_id)
    except ValueError:
        raise HTTPException(400, "Невалидный id")
    plan = await _get_plan_owned(db, pid, current_user.company_id)

    today = date.today()
    cabinet = (uuid.UUID(plan.scope_ref)
               if plan.scope_type == "cabinet" and plan.scope_ref else None)
    extra = "AND oa.id = :cab" if cabinet else ""
    base_params = {"cid": str(current_user.company_id)}
    if cabinet:
        base_params["cab"] = str(cabinet)

    # Период плана
    days_total = (plan.period_end - plan.period_start).days + 1
    plan_target = float(plan.target_value)

    async def fetch(df: date, dt: date) -> dict:
        params = {**base_params, "df": df, "dt": dt}
        sales_q = f"""
            SELECT
              COUNT(DISTINCT o.id)::float AS orders,
              COALESCE(SUM(oi.quantity), 0)::float AS units
            FROM orders o
            JOIN order_items oi ON oi.order_id = o.id
            JOIN ozon_accounts oa ON oa.id = o.ozon_account_id
            WHERE oa.company_id = :cid AND o.status='delivered'
              AND o.order_created_at >= :df AND o.order_created_at <= :dt
              {extra}
        """
        sales = (await db.execute(_sql(sales_q), params)).first()

        fin_q = f"""
            SELECT
              COALESCE(SUM(t.accruals_for_sale) FILTER (
                WHERE t.operation_type='OperationAgentDeliveredToCustomer'), 0)::float AS revenue,
              COALESCE(SUM(ABS(t.sale_commission)), 0)::float AS commission,
              COALESCE(SUM(ABS(t.delivery_to_customer)), 0)::float AS logistics,
              COALESCE(SUM(ABS(t.acquiring)), 0)::float AS acquiring,
              COALESCE(SUM(ABS(t.advertising)), 0)::float AS advertising
            FROM transactions t
            JOIN ozon_accounts oa ON oa.id = t.ozon_account_id
            WHERE oa.company_id = :cid
              AND t.operation_date >= :df AND t.operation_date <= :dt
              {extra}
        """
        fin = (await db.execute(_sql(fin_q), params)).first()

        rev = float(fin.revenue or 0)
        ad = float(fin.advertising or 0)
        marginal = rev - float(fin.commission or 0) - float(fin.logistics or 0) \
                       - float(fin.acquiring or 0) - ad
        # Чистая прибыль: эвристика УСН 6% (точнее — через services/tax.py, но
        # для сводки достаточно проксики)
        net = marginal - rev * 0.06

        drr = (ad / rev * 100) if rev > 0 else 0

        return {
            "orders": float(sales.orders or 0),
            "units": float(sales.units or 0),
            "revenue": rev,
            "marginal_profit": marginal,
            "net_profit": net,
            "ad_spend": ad,
            "drr_pct": drr,
        }

    # 3 среза
    today_data = await fetch(today, today)
    cum_data = await fetch(plan.period_start, min(today, plan.period_end))
    month_start = date(today.year, today.month, 1)
    if today.month == 12:
        month_end = date(today.year + 1, 1, 1) - _td(days=1)
    else:
        month_end = date(today.year, today.month + 1, 1) - _td(days=1)
    month_data = await fetch(month_start, min(today, month_end))

    days_elapsed = max(0, min(days_total, (today - plan.period_start).days + 1))

    def prorata(metric_target_total: float) -> float:
        return metric_target_total * days_elapsed / days_total if days_total else 0

    # Pro-rata план на разные срезы
    plan_per_day = plan_target / days_total if days_total else 0
    plan_today = plan_per_day if (plan.period_start <= today <= plan.period_end) else 0
    plan_cum = prorata(plan_target)
    # Месяц pro-rata: пересечение плана и месяца / план_per_day × дни_пересечения
    overlap_start = max(month_start, plan.period_start)
    overlap_end = min(month_end, plan.period_end, today)
    overlap_days = max(0, (overlap_end - overlap_start).days + 1)
    plan_month = plan_per_day * overlap_days

    def make_cell(fact_v: float, plan_v: float) -> dict:
        pct = (fact_v / plan_v * 100) if plan_v > 0 else None
        return {
            "fact": round(fact_v, 2),
            "plan": round(plan_v, 2),
            "pct": round(pct, 1) if pct is not None else None,
        }

    # Метрики смапим: план есть только для plan.metric_code, остальные — без плана
    def build_group(metric_key: str, label: str, group: str) -> dict:
        # Только для main metric плана есть pro-rata план
        is_main = metric_key == plan.metric_code
        plan_t = plan_today if is_main else 0
        plan_c = plan_cum if is_main else 0
        plan_m = plan_month if is_main else 0
        return {
            "code": metric_key, "label": label, "group": group,
            "is_main": is_main,
            "today": make_cell(today_data.get(metric_key, 0), plan_t),
            "cumulative": make_cell(cum_data.get(metric_key, 0), plan_c),
            "month": make_cell(month_data.get(metric_key, 0), plan_m),
        }

    return {
        "plan_id": str(pid),
        "plan_name": plan.name,
        "plan_metric": plan.metric_code,
        "days_elapsed": days_elapsed,
        "days_total": days_total,
        "groups": {
            "ПРОДАЖИ": [
                build_group("orders", "Заказы", "ПРОДАЖИ"),
                build_group("units", "Единицы", "ПРОДАЖИ"),
                build_group("revenue", "Выручка", "ПРОДАЖИ"),
            ],
            "ФИНАНСЫ": [
                build_group("marginal_profit", "Маржинальная прибыль", "ФИНАНСЫ"),
                build_group("net_profit", "Чистая прибыль (≈УСН 6%)", "ФИНАНСЫ"),
            ],
            "ПРОДВИЖЕНИЕ": [
                build_group("ad_spend", "Расход на рекламу", "ПРОДВИЖЕНИЕ"),
                build_group("drr_pct", "ДРР, %", "ПРОДВИЖЕНИЕ"),
            ],
        },
    }


@router.get("/{plan_id}/streak")
async def green_streak(
    plan_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Серия дней «в зелёной зоне» (pro-rata план выполнен на день).

    Идём от сегодня назад: пока факт-день ≥ план-день — серия растёт.
    """
    from sqlalchemy import text as _sql
    from datetime import timedelta as _td
    try:
        pid = uuid.UUID(plan_id)
    except ValueError:
        raise HTTPException(400, "Невалидный id")
    plan = await _get_plan_owned(db, pid, current_user.company_id)

    today = date.today()
    days_total = (plan.period_end - plan.period_start).days + 1
    plan_per_day = float(plan.target_value) / days_total if days_total else 0

    cabinet = (uuid.UUID(plan.scope_ref)
               if plan.scope_type == "cabinet" and plan.scope_ref else None)
    extra = "AND oa.id = :cab" if cabinet else ""
    params = {"cid": str(current_user.company_id),
              "df": plan.period_start, "dt": min(today, plan.period_end)}
    if cabinet:
        params["cab"] = str(cabinet)

    if plan.metric_code == "revenue":
        sql = f"""
            SELECT t.operation_date::date AS d,
                   COALESCE(SUM(t.accruals_for_sale), 0)::float AS v
            FROM transactions t
            JOIN ozon_accounts oa ON oa.id = t.ozon_account_id
            WHERE oa.company_id = :cid
              AND t.operation_date >= :df AND t.operation_date <= :dt
              AND t.operation_type='OperationAgentDeliveredToCustomer' {extra}
            GROUP BY 1 ORDER BY 1 DESC
        """
    elif plan.metric_code == "orders":
        sql = f"""
            SELECT o.created_at::date AS d, COUNT(*)::float AS v
            FROM orders o
            JOIN ozon_accounts oa ON oa.id = o.ozon_account_id
            WHERE oa.company_id = :cid AND o.status='delivered'
              AND o.order_created_at >= :df AND o.order_created_at <= :dt {extra}
            GROUP BY 1 ORDER BY 1 DESC
        """
    else:
        return {"streak": 0, "note": "Серия считается для revenue и orders"}

    rows = (await db.execute(_sql(sql), params)).all()
    by_day = {r.d: float(r.v or 0) for r in rows}

    # Идём от сегодня назад
    streak = 0
    best_streak = 0
    cur = 0
    d = min(today, plan.period_end)
    while d >= plan.period_start:
        v = by_day.get(d, 0)
        if v >= plan_per_day:
            cur += 1
            best_streak = max(best_streak, cur)
            if d == min(today, plan.period_end) or streak == cur - 1:
                # Текущая серия (от сегодня без пропусков)
                streak = cur
        else:
            if cur > best_streak:
                best_streak = cur
            cur = 0
        d -= _td(days=1)

    return {
        "streak": streak,
        "best_streak": best_streak,
        "plan_per_day": round(plan_per_day, 2),
        "note": f"Серия дней где факт ≥ {plan_per_day:.0f}/день",
    }


@router.get("/{plan_id}/timeseries")
async def fact_timeseries(
    plan_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Burn-up: дневной факт + дневной план (pro-rata) для графика.

    Возвращает 3 ряда:
      [{day, fact_cum, plan_cum, run_rate_cum}, ...]
    """
    from datetime import timedelta
    try:
        pid = uuid.UUID(plan_id)
    except ValueError:
        raise HTTPException(400, "Невалидный id")
    plan = await _get_plan_owned(db, pid, current_user.company_id)

    today = date.today()
    days_total = (plan.period_end - plan.period_start).days + 1
    plan_total = float(plan.target_value)
    daily_plan = plan_total / days_total

    cabinet = (uuid.UUID(plan.scope_ref)
               if plan.scope_type == "cabinet" and plan.scope_ref else None)
    extra = "AND oa.id = :cab" if cabinet else ""
    params = {"cid": str(current_user.company_id),
              "df": plan.period_start, "dt": plan.period_end}
    if cabinet:
        params["cab"] = str(cabinet)

    if plan.metric_code == "revenue":
        sql = f"""
            SELECT t.operation_date::date AS d,
                   COALESCE(SUM(t.accruals_for_sale), 0)::float AS v
            FROM transactions t
            JOIN ozon_accounts oa ON oa.id = t.ozon_account_id
            WHERE oa.company_id = :cid
              AND t.operation_date >= :df AND t.operation_date <= :dt
              AND t.operation_type='OperationAgentDeliveredToCustomer'
              {extra}
            GROUP BY 1 ORDER BY 1
        """
    elif plan.metric_code == "orders":
        sql = f"""
            SELECT o.created_at::date AS d, COUNT(*)::float AS v
            FROM orders o
            JOIN ozon_accounts oa ON oa.id = o.ozon_account_id
            WHERE oa.company_id = :cid AND o.status='delivered'
              AND o.order_created_at >= :df AND o.order_created_at <= :dt
              {extra}
            GROUP BY 1 ORDER BY 1
        """
    else:
        sql = f"""
            SELECT t.operation_date::date AS d,
                   COALESCE(SUM(t.accruals_for_sale), 0)::float AS v
            FROM transactions t
            JOIN ozon_accounts oa ON oa.id = t.ozon_account_id
            WHERE oa.company_id = :cid
              AND t.operation_date >= :df AND t.operation_date <= :dt
              {extra}
            GROUP BY 1 ORDER BY 1
        """
    from sqlalchemy import text as _sql
    fact_rows = (await db.execute(_sql(sql), params)).all()
    fact_by_day = {r.d: float(r.v or 0) for r in fact_rows}

    series = []
    fact_cum = 0.0
    plan_cum = 0.0
    fact_so_far = 0.0
    days_elapsed = 0
    for i in range(days_total):
        d = plan.period_start + timedelta(days=i)
        plan_cum += daily_plan
        if d <= today:
            fact_cum += fact_by_day.get(d, 0)
            fact_so_far = fact_cum
            days_elapsed = i + 1
        series.append({
            "day": d.isoformat(),
            "fact_cum": round(fact_cum, 2) if d <= today else None,
            "plan_cum": round(plan_cum, 2),
        })

    # Run-rate проекция: продолжаем средним темпом
    if days_elapsed > 0 and days_elapsed < days_total:
        avg_daily = fact_so_far / days_elapsed
        for i in range(days_elapsed, days_total):
            series[i]["run_rate_cum"] = round(
                fact_so_far + avg_daily * (i - days_elapsed + 1), 2
            )

    return {"series": series, "plan_total": plan_total, "today": today.isoformat()}


@router.patch("/{plan_id}/kpi/{kpi_id}")
async def update_kpi(
    plan_id: str,
    kpi_id: str,
    payload: KPICreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    try:
        pid = uuid.UUID(plan_id)
        kid = uuid.UUID(kpi_id)
    except ValueError:
        raise HTTPException(400, "Невалидный id")
    await _get_plan_owned(db, pid, current_user.company_id)
    kpi = (await db.execute(
        select(PlanKPI).where(PlanKPI.id == kid, PlanKPI.plan_id == pid)
    )).scalar_one_or_none()
    if not kpi:
        raise HTTPException(404, "KPI не найден")
    kpi.manager_name = payload.manager_name
    kpi.metric_code = payload.metric_code
    kpi.target_value = Decimal(str(payload.target_value))
    kpi.bonus_rule = payload.bonus_rule
    await db.commit()
    return {"id": str(kpi.id), "ok": True}


@router.delete("/{plan_id}/kpi/{kpi_id}")
async def delete_kpi(
    plan_id: str,
    kpi_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    try:
        pid = uuid.UUID(plan_id)
        kid = uuid.UUID(kpi_id)
    except ValueError:
        raise HTTPException(400, "Невалидный id")
    await _get_plan_owned(db, pid, current_user.company_id)
    kpi = (await db.execute(
        select(PlanKPI).where(PlanKPI.id == kid, PlanKPI.plan_id == pid)
    )).scalar_one_or_none()
    if not kpi:
        raise HTTPException(404, "KPI не найден")
    await db.delete(kpi)
    await db.commit()
    return {"ok": True}


@router.post("/{plan_id}/kpi")
async def create_kpi(
    plan_id: str,
    payload: KPICreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    try:
        pid = uuid.UUID(plan_id)
    except ValueError:
        raise HTTPException(400, "Невалидный id")
    await _get_plan_owned(db, pid, current_user.company_id)
    kpi = PlanKPI(
        plan_id=pid, manager_name=payload.manager_name,
        metric_code=payload.metric_code,
        target_value=Decimal(str(payload.target_value)),
        bonus_rule=payload.bonus_rule,
    )
    db.add(kpi)
    await db.commit()
    return {
        "id": str(kpi.id), "manager_name": kpi.manager_name,
        "metric_code": kpi.metric_code, "target_value": float(kpi.target_value),
        "bonus_rule": kpi.bonus_rule,
    }


@router.get("/{plan_id}/kpi")
async def list_kpi(
    plan_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    try:
        pid = uuid.UUID(plan_id)
    except ValueError:
        raise HTTPException(400, "Невалидный id")
    await _get_plan_owned(db, pid, current_user.company_id)
    rows = (await db.execute(
        select(PlanKPI).where(PlanKPI.plan_id == pid).order_by(PlanKPI.created_at)
    )).scalars().all()
    return [
        {
            "id": str(k.id), "manager_name": k.manager_name,
            "metric_code": k.metric_code,
            "target_value": float(k.target_value),
            "bonus_rule": k.bonus_rule,
        }
        for k in rows
    ]
